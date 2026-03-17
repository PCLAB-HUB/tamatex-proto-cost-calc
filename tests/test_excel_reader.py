# テスト依存: pytest, openpyxl
"""excel_reader モジュールの単体テスト。

_convert_cell_value() と read_workbook() を検証する。
実際の xlsx ファイルは openpyxl で一時ディレクトリに生成して使用する。
"""

from datetime import date, datetime, time

import openpyxl
import pytest

from tamatex.excel_reader import (
    SheetData,
    WorkbookData,
    _convert_cell_value,
    read_workbook,
)


# ---------------------------------------------------------------------------
# _convert_cell_value テスト
# ---------------------------------------------------------------------------

def test_convert_cell_value_none_returns_empty_string():
    """None は空文字列に変換されること。"""
    assert _convert_cell_value(None) == ""


def test_convert_cell_value_datetime_returns_formatted_string():
    """datetime は 'YYYY-MM-DD HH:MM:SS' 形式の文字列に変換されること。"""
    dt = datetime(2024, 3, 15, 9, 30, 0)
    result = _convert_cell_value(dt)
    assert result == "2024-03-15 09:30:00"


def test_convert_cell_value_date_returns_formatted_string():
    """date は 'YYYY-MM-DD' 形式の文字列に変換されること。"""
    d = date(2024, 12, 31)
    result = _convert_cell_value(d)
    assert result == "2024-12-31"


def test_convert_cell_value_time_returns_formatted_string():
    """time は 'HH:MM:SS' 形式の文字列に変換されること。"""
    t = time(8, 5, 3)
    result = _convert_cell_value(t)
    assert result == "08:05:03"


def test_convert_cell_value_bool_true_returns_true():
    """bool の True はそのまま True として返されること。"""
    result = _convert_cell_value(True)
    assert result is True


def test_convert_cell_value_bool_false_returns_false():
    """bool の False はそのまま False として返されること。"""
    result = _convert_cell_value(False)
    assert result is False


def test_convert_cell_value_bool_is_not_converted_to_int():
    """bool 値は int に変換されないこと（bool は int のサブクラスなので順序が重要）。"""
    # True が 1 に変換されてしまうと bool チェックが先行していない証拠
    result = _convert_cell_value(True)
    assert isinstance(result, bool)


def test_convert_cell_value_int_returns_int():
    """int 値はそのまま int として返されること。"""
    result = _convert_cell_value(42)
    assert result == 42
    assert isinstance(result, int)


def test_convert_cell_value_float_returns_float():
    """float 値はそのまま float として返されること。"""
    result = _convert_cell_value(3.14)
    assert result == pytest.approx(3.14)
    assert isinstance(result, float)


def test_convert_cell_value_string_returns_string():
    """str 値はそのまま返されること。"""
    result = _convert_cell_value("hello")
    assert result == "hello"


def test_convert_cell_value_arbitrary_object_returns_str():
    """その他の型は str() で文字列化されること。"""

    class Custom:
        def __str__(self):
            return "custom_value"

    result = _convert_cell_value(Custom())
    assert result == "custom_value"


def test_convert_cell_value_zero_int_returns_zero():
    """0 (int) は空文字ではなく 0 として返されること。"""
    result = _convert_cell_value(0)
    assert result == 0


def test_convert_cell_value_zero_float_returns_zero():
    """0.0 (float) は空文字ではなく 0.0 として返されること。"""
    result = _convert_cell_value(0.0)
    assert result == 0.0


# ---------------------------------------------------------------------------
# ヘルパー: テスト用 xlsx ファイル生成
# ---------------------------------------------------------------------------

def _create_simple_xlsx(path, sheet_name: str, data: list[list]) -> None:
    """シンプルな xlsx ファイルを一時パスに書き出す。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in data:
        ws.append(row)
    wb.save(str(path))
    wb.close()


# ---------------------------------------------------------------------------
# read_workbook テスト
# ---------------------------------------------------------------------------

def test_read_workbook_returns_workbook_data(tmp_path):
    """read_workbook が WorkbookData を返すこと。"""
    xlsx_path = tmp_path / "test.xlsx"
    _create_simple_xlsx(xlsx_path, "Sheet1", [["A", "B"], [1, 2]])

    result = read_workbook(xlsx_path)
    assert isinstance(result, WorkbookData)


def test_read_workbook_file_name_is_stem(tmp_path):
    """WorkbookData.file_name がファイルの stem（拡張子なし名）であること。"""
    xlsx_path = tmp_path / "my_report.xlsx"
    _create_simple_xlsx(xlsx_path, "Sheet1", [["x"]])

    result = read_workbook(xlsx_path)
    assert result.file_name == "my_report"


def test_read_workbook_sheets_is_list_of_sheet_data(tmp_path):
    """WorkbookData.sheets が SheetData のリストであること。"""
    xlsx_path = tmp_path / "data.xlsx"
    _create_simple_xlsx(xlsx_path, "Sheet1", [["value"]])

    result = read_workbook(xlsx_path)
    assert all(isinstance(s, SheetData) for s in result.sheets)


def test_read_workbook_reads_sheet_name(tmp_path):
    """シート名が正しく読み取られること。"""
    xlsx_path = tmp_path / "named.xlsx"
    _create_simple_xlsx(xlsx_path, "売上データ", [["col1", "col2"]])

    result = read_workbook(xlsx_path)
    assert result.sheets[0].name == "売上データ"


def test_read_workbook_reads_cell_values(tmp_path):
    """セルの値が正しく読み取られること。"""
    xlsx_path = tmp_path / "values.xlsx"
    _create_simple_xlsx(xlsx_path, "Sheet1", [["名前", "金額"], ["田中", 1000]])

    result = read_workbook(xlsx_path)
    rows = result.sheets[0].rows

    assert rows[0][0] == "名前"
    assert rows[0][1] == "金額"
    assert rows[1][0] == "田中"
    assert rows[1][1] == 1000


def test_read_workbook_removes_trailing_empty_rows(tmp_path):
    """末尾の空行が除去されること。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["data", "row"])
    ws.append([None, None])   # 空行
    ws.append([None, None])   # 空行
    xlsx_path = tmp_path / "trailing_empty.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = read_workbook(xlsx_path)
    rows = result.sheets[0].rows

    # 末尾空行が除去されてデータ行のみ残ること
    assert len(rows) == 1
    assert rows[0][0] == "data"


def test_read_workbook_keeps_empty_rows_in_middle(tmp_path):
    """中間の空行は除去されないこと。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["row1"])
    ws.append([None])    # 中間の空行
    ws.append(["row3"])
    xlsx_path = tmp_path / "middle_empty.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = read_workbook(xlsx_path)
    rows = result.sheets[0].rows

    assert len(rows) == 3


def test_read_workbook_multiple_sheets(tmp_path):
    """複数シートを持つファイルの全シートが読み取られること。"""
    wb = openpyxl.Workbook()
    wb.active.title = "シート1"
    wb.active.append(["a", "b"])
    ws2 = wb.create_sheet("シート2")
    ws2.append(["c", "d"])
    xlsx_path = tmp_path / "multi_sheet.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = read_workbook(xlsx_path)

    assert len(result.sheets) == 2
    sheet_names = [s.name for s in result.sheets]
    assert "シート1" in sheet_names
    assert "シート2" in sheet_names


def test_read_workbook_empty_sheet_returns_zero_rows(tmp_path):
    """空シートは rows が空リストであること。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    xlsx_path = tmp_path / "empty_sheet.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = read_workbook(xlsx_path)
    assert result.sheets[0].rows == []


def test_read_workbook_accepts_path_object(tmp_path):
    """Path オブジェクトを渡しても正常に動作すること。"""
    from pathlib import Path
    xlsx_path = tmp_path / "path_obj.xlsx"
    _create_simple_xlsx(xlsx_path, "Sheet1", [["ok"]])

    result = read_workbook(Path(xlsx_path))
    assert isinstance(result, WorkbookData)


def test_read_workbook_accepts_string_path(tmp_path):
    """文字列パスを渡しても正常に動作すること。"""
    xlsx_path = tmp_path / "str_path.xlsx"
    _create_simple_xlsx(xlsx_path, "Sheet1", [["ok"]])

    result = read_workbook(str(xlsx_path))
    assert isinstance(result, WorkbookData)


def test_read_workbook_all_empty_rows_removed(tmp_path):
    """全行が空の場合、rows が空リストになること。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AllEmpty"
    ws.append([None, None, None])
    ws.append([None, None, None])
    xlsx_path = tmp_path / "all_empty.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = read_workbook(xlsx_path)
    assert result.sheets[0].rows == []
