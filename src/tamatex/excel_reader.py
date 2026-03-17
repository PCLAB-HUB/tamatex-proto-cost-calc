"""Excel読み取りモジュール。openpyxlでExcelデータをPythonデータ構造に変換する。"""

import logging
from dataclasses import dataclass
from datetime import datetime, date, time
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger("tamatex")


@dataclass
class SheetData:
    name: str
    rows: list[list]  # 各行は値のリスト


@dataclass
class WorkbookData:
    file_name: str
    sheets: list[SheetData]


def _convert_cell_value(value) -> str | int | float | None:
    """セル値をGoogle Sheets互換の型に変換する。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _expand_merged_cells(ws, rows: list[list]) -> None:
    """結合セルの値を結合範囲全体に展開する（in-place）。

    read_only=False モードでのみ ws.merged_cells.ranges が利用可能。
    結合範囲の左上セルの値を、その範囲内の全セルにコピーする。
    """
    for merged_range in ws.merged_cells.ranges:
        # 結合範囲の左上セルの値を取得（0-indexed に変換）
        top_row = merged_range.min_row - 1
        left_col = merged_range.min_col - 1

        if top_row >= len(rows):
            continue
        if left_col >= len(rows[top_row]):
            continue

        anchor_value = rows[top_row][left_col]

        # 結合範囲内の全セルに左上の値を展開
        for r in range(merged_range.min_row - 1, merged_range.max_row):
            for c in range(merged_range.min_col - 1, merged_range.max_col):
                if r < len(rows) and c < len(rows[r]):
                    rows[r][c] = anchor_value


def read_workbook(file_path: str | Path) -> WorkbookData:
    """Excelファイルを読み取り、全シートのデータを返す。

    結合セルは左上セルの値で全体を埋める。
    数式セルは計算済み値（data_only=True）を使用する。
    """
    path = Path(file_path)
    logger.info("Excel読み取り開始: %s", path.name)

    # read_only=False にすることで merged_cells.ranges が利用可能になる
    wb = load_workbook(str(path), read_only=False, data_only=True)
    try:
        sheets: list[SheetData] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list] = []
            for row in ws.iter_rows():
                converted = [_convert_cell_value(cell.value) for cell in row]
                rows.append(converted)

            # 結合セルの値を展開
            _expand_merged_cells(ws, rows)

            # 末尾の空行を除去
            while rows and all(v == "" or v is None for v in rows[-1]):
                rows.pop()

            sheets.append(SheetData(name=sheet_name, rows=rows))
            logger.debug("  シート '%s': %d 行", sheet_name, len(rows))

        return WorkbookData(file_name=path.stem, sheets=sheets)
    finally:
        wb.close()
