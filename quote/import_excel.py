"""原価計算書参考資料.xlsx の全商品データをツールにインポートする."""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from quote.data.mock_data import seed_mock_data
from quote.engine.models import ContainerExpenses, GlobalParams, ProductInput
from quote.storage.db import init_db, save_quote

EXCEL_PATH = Path(__file__).resolve().parent.parent / "原価計算書参考資料.xlsx"
if not EXCEL_PATH.exists():
    EXCEL_PATH = Path("/Users/pclab/Desktop/Project/tamatex/原価計算書参考資料.xlsx")


def _col(letter):
    return column_index_from_string(letter)


def _val(ws, row, col_letter, default=None):
    v = ws.cell(row=row, column=_col(col_letter)).value
    return v if v is not None else default


def _float(ws, row, col_letter, default=0.0):
    v = ws.cell(row=row, column=_col(col_letter)).value
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _int(ws, row, col_letter, default=0):
    v = ws.cell(row=row, column=_col(col_letter)).value
    if v is None:
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def import_all():
    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active

    seed_mock_data()
    init_db()

    groups = {}
    for row in range(9, 134):
        name = _val(ws, row, 'H')
        if not name:
            continue

        weight_g = _float(ws, row, 'M')
        momme = weight_g / 3.75 if weight_g > 0 else 0.0

        p = ProductInput(
            supplier=_val(ws, row, 'A', ''),
            inspection_pass=(_val(ws, row, 'B', '') == '○'),
            customer=_val(ws, row, 'C', ''),
            port=_val(ws, row, 'D', ''),
            delivery_to=_val(ws, row, 'E', ''),
            container_ft=_int(ws, row, 'F', 20),
            ship_to=_val(ws, row, 'G', ''),
            product_name=str(name),
            prototype_code=str(_val(ws, row, 'I', '')),
            item_type=_val(ws, row, 'J', ''),
            package_size_cm=_val(ws, row, 'K', ''),
            weight_momme=momme,
            weight_g=weight_g,
            fabric_quality=_val(ws, row, 'N', ''),
            method=_val(ws, row, 'O', ''),
            packing_quantity=_int(ws, row, 'P', 1),
            packing_size=_val(ws, row, 'Q', ''),
            container_load=_float(ws, row, 'R'),
            fob_usd=_float(ws, row, 'S'),
            other_processing_usd=_float(ws, row, 'T'),
            loss_rate=_float(ws, row, 'U'),
            charge_up_unit=_float(ws, row, 'V'),
            embroidery_per_1000=_float(ws, row, 'W', 0.03),
            stitch_count=_float(ws, row, 'X'),
            die_charge=_float(ws, row, 'Y'),
            inspection_jpy=_float(ws, row, 'AA'),
            packing_jpy=_float(ws, row, 'AB'),
            material_jpy=_float(ws, row, 'AC'),
            inspection_cny=_float(ws, row, 'AD'),
            packing_cny=_float(ws, row, 'AE'),
            material_cny=_float(ws, row, 'AF'),
            inspection_usd=_float(ws, row, 'AG'),
            packing_usd=_float(ws, row, 'AH'),
            material_usd=_float(ws, row, 'AI'),
            tariff_rate_override=_float(ws, row, 'AP'),
            quote_price_ex_amort=_float(ws, row, 'DZ'),
            center_fee_ex_amort=_float(ws, row, 'EA'),
            rebate_ex_amort=_float(ws, row, 'EB'),
            retail_price_ex_amort=_float(ws, row, 'EI'),
            ribbon=_float(ws, row, 'BH'),
            name_label_2=_float(ws, row, 'BI'),
            name_label_3=_float(ws, row, 'BJ'),
            seal_1=_float(ws, row, 'BK'),
            seal_2=_float(ws, row, 'BL'),
            tag=_float(ws, row, 'BM'),
            bag=_float(ws, row, 'BN'),
            other_material=_float(ws, row, 'BO'),
            material_freight=_float(ws, row, 'BP'),
            design_cost=_float(ws, row, 'BU'),
            jq_card=_float(ws, row, 'BV'),
            embroidery_card=_float(ws, row, 'BW'),
            print_unit_price=_float(ws, row, 'BX'),
            print_type_count=_float(ws, row, 'BY'),
            layout=_float(ws, row, 'CA'),
            name_plate=_float(ws, row, 'CB'),
            seal_plate=_float(ws, row, 'CC'),
            tab_plate=_float(ws, row, 'CD'),
            bag_plate=_float(ws, row, 'CE'),
            cardboard_plate=_float(ws, row, 'CF'),
            other_depreciation=_float(ws, row, 'CG'),
            sample_cost=_float(ws, row, 'CH'),
            quality_inspection=_float(ws, row, 'CI'),
            other_amortization=_float(ws, row, 'CJ'),
            logistics_cardboard=_float(ws, row, 'CN'),
            logistics_io_fee=_float(ws, row, 'CO'),
            logistics_storage_months=_float(ws, row, 'CP'),
            logistics_storage_fee=_float(ws, row, 'CQ'),
            logistics_slip_fee=_float(ws, row, 'CR'),
            logistics_freight=_float(ws, row, 'CS'),
            domestic_packing_qty=_int(ws, row, 'CV', 1),
            domestic_processing=_float(ws, row, 'CW'),
            domestic_material=_float(ws, row, 'CX'),
            domestic_cardboard=_float(ws, row, 'CY'),
            domestic_io=_float(ws, row, 'CZ'),
            domestic_storage_months=_float(ws, row, 'DA'),
            domestic_storage_fee=_float(ws, row, 'DB'),
            domestic_freight=_float(ws, row, 'DC'),
            quote_price=_float(ws, row, 'DH'),
            center_fee=_float(ws, row, 'DI'),
            rebate=_float(ws, row, 'DJ'),
            lot_per_color=_int(ws, row, 'DK'),
            num_colors=_int(ws, row, 'DL', 1),
            retail_price=_float(ws, row, 'DP'),
        )
        rate = _float(ws, row, 'AJ', 152.0)
        current_rate = _float(ws, row, 'AK', rate)
        container_key = (
            rate,
            current_rate,
            _float(ws, row, 'AS'), _float(ws, row, 'AT'),
            _float(ws, row, 'AU'), _float(ws, row, 'AV'),
            _float(ws, row, 'AW'), _float(ws, row, 'AX'),
            _float(ws, row, 'AY'), _float(ws, row, 'AZ'),
            _float(ws, row, 'BA'), _float(ws, row, 'BB'),
            _float(ws, row, 'BC'),
        )

        if container_key not in groups:
            groups[container_key] = []
        groups[container_key].append(p)
        print(f"  Row {row}: {p.product_name} (FOB=${p.fob_usd}, 内部為替={rate}, 現行為替={current_rate})")

    print(f"\n=== 為替・経費グループ: {len(groups)}件 ===")
    for i, (key, items) in enumerate(groups.items()):
        rate = key[0]
        current_rate = key[1]
        ce_vals = key[2:]

        params = GlobalParams(
            internal_rate=rate,
            current_rate=current_rate,
            overseas_freight_usd=240.0,
            cny_to_usd_rate=0.17,
            cny_to_jpy_rate=13.0,
            insurance_risk_rate=0.0018,
            tariff_rate=0.0,
            b_grade_loss_rate=0.01,
            sub_material_loss_rate=0.05,
            amortization_margin=0.05,
            margin=0.20,
            container_expenses=ContainerExpenses(
                cy_charge=ce_vals[0],
                lss=ce_vals[1],
                lss_cic_usd=ce_vals[2],
                thc=ce_vals[3],
                emc=ce_vals[4],
                do_fee=ce_vals[5],
                doc_fee=ce_vals[6],
                customs_fee=ce_vals[7],
                handling_fee=ce_vals[8],
                drayage=ce_vals[9],
                devanning=ce_vals[10],
            ),
        )

        title = f"Excel検証 為替{rate:.0f}円 ({len(items)}商品)"
        quote_id = save_quote(
            customer_id=1,
            staff_id=1,
            title=title,
            products=items,
            params=params,
            notes=f"為替={rate}円, コンテナ経費計={sum(ce_vals[0:2]) + ce_vals[2]*rate + sum(ce_vals[3:]):,.0f}",
        )
        print(f"  グループ{i+1}: 為替{rate:.0f}円, {len(items)}商品 → 見積もりID={quote_id}")

    total = sum(len(v) for v in groups.values())
    print(f"\n=== インポート完了 ===")
    print(f"総商品数: {total}")
    print(f"見積もり数: {len(groups)}")


if __name__ == "__main__":
    import_all()
